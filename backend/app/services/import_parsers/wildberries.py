"""
Wildberries export file parser.
Handles XLSX files exported from Wildberries seller portal.
"""

from typing import Any, Iterator

from openpyxl import load_workbook

from .base import BaseImportParser


class WildberriesParser(BaseImportParser):
    """Parser for Wildberries XLSX export files."""

    source_name = "Wildberries"
    supported_extensions = ['.xlsx']

    # Known Wildberries column names (may vary by export type)
    KNOWN_COLUMNS = {
        # Product identification
        'Артикул поставщика': 'sku',
        'Артикул продавца': 'sku',
        'Артикул WB': 'barcode',
        'Баркод товара': 'barcode',
        'Штрихкод': 'barcode',

        # Product info
        'Наименование': 'name',
        'Название товара': 'name',
        'Бренд': 'tags',
        'Предмет': 'category',
        'Категория': 'category',
        'Описание': 'long_description',

        # Pricing
        'Цена': 'price',
        'Цена со скидкой': 'price',
        'Розничная цена': 'price',
        'Цена до скидки': 'original_price',

        # Stock
        'Остаток': 'stock_quantity',
        'Доступно к заказу': 'stock_quantity',
        'Количество': 'stock_quantity',

        # Images
        'Фото': 'main_image_url',
        'Медиафайлы': 'gallery_urls',
        'Ссылка на фото': 'main_image_url',
        'Изображения': 'gallery_urls',

        # Variants
        'Размер': 'size',
        'Цвет': 'color',
        'Российский размер': 'size',
    }

    def get_columns(self, file_path: str) -> list[str]:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Wildberries may have header in first or second row
            for row_num in range(1, 4):
                row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True), None)
                if row and any(cell for cell in row):
                    headers = [str(cell).strip() if cell else '' for cell in row]
                    # Check if this looks like a header row
                    if any(h in self.KNOWN_COLUMNS for h in headers if h):
                        wb.close()
                        return headers

            # Fallback to first row
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()

            if first_row:
                return [str(cell).strip() if cell else f'Column_{i}'
                        for i, cell in enumerate(first_row)]
            return []
        except Exception:
            return []

    def iter_rows(self, file_path: str) -> Iterator[dict[str, Any]]:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Find header row
            header_row = 1
            for row_num in range(1, 4):
                row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True), None)
                if row:
                    headers = [str(cell).strip() if cell else '' for cell in row]
                    if any(h in self.KNOWN_COLUMNS for h in headers if h):
                        header_row = row_num
                        break

            # Get headers
            headers = []
            for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
                headers = [str(cell).strip() if cell else f'Column_{i}'
                          for i, cell in enumerate(row)]
                break

            if not headers:
                wb.close()
                return

            # Iterate data rows
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if row and any(cell is not None for cell in row):
                    row_data = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            value = self.normalize_value(cell)
                            row_data[headers[i]] = value
                    yield row_data

            wb.close()
        except Exception:
            return

    def get_suggested_mapping(self) -> dict[str, str]:
        """Return Wildberries-specific column mappings."""
        return {k.lower(): v for k, v in self.KNOWN_COLUMNS.items()}

    def parse_variant_attributes(self, row: dict[str, Any]) -> dict[str, str]:
        """
        Extract variant attributes from a Wildberries row.

        Returns:
            Dictionary of attribute_name -> attribute_value
        """
        attributes = {}

        # Common variant fields
        variant_fields = [
            ('Размер', 'Размер'),
            ('Российский размер', 'Размер'),
            ('Цвет', 'Цвет'),
            ('Рост', 'Рост'),
            ('Длина', 'Длина'),
            ('Ширина', 'Ширина'),
        ]

        for col_name, attr_name in variant_fields:
            value = row.get(col_name)
            if value and str(value).strip():
                attributes[attr_name] = str(value).strip()

        return attributes

    def extract_image_urls(self, row: dict[str, Any]) -> tuple[str | None, list[str]]:
        """
        Extract main image and gallery URLs from a row.

        Returns:
            Tuple of (main_image_url, gallery_urls)
        """
        main_image = None
        gallery = []

        # Check main image columns
        for col in ['Фото', 'Ссылка на фото', 'Изображение']:
            value = row.get(col)
            if value:
                urls = self.parse_images(value)
                if urls:
                    main_image = urls[0]
                    gallery.extend(urls[1:])
                    break

        # Check gallery columns
        for col in ['Медиафайлы', 'Изображения', 'Дополнительные фото']:
            value = row.get(col)
            if value:
                urls = self.parse_images(value)
                gallery.extend(urls)

        return main_image, list(set(gallery))  # Remove duplicates

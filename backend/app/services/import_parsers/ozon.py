"""
Ozon export file parser.
Handles XLS/CSV files exported from Ozon seller portal.
"""

import csv
from typing import Any, Iterator

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

from openpyxl import load_workbook

from .base import BaseImportParser


class OzonParser(BaseImportParser):
    """Parser for Ozon XLS/XLSX/CSV export files."""

    source_name = "Ozon"
    supported_extensions = ['.xls', '.xlsx', '.csv']

    # Known Ozon column names
    KNOWN_COLUMNS = {
        # Product identification
        'offer_id': 'sku',
        'Артикул': 'sku',
        'id': 'external_id',
        'sku': 'sku',
        'barcode': 'barcode',
        'Штрихкод': 'barcode',

        # Product info
        'name': 'name',
        'Название': 'name',
        'Наименование товара': 'name',
        'description': 'long_description',
        'Описание': 'long_description',
        'category_id': 'category_id',
        'Категория': 'category',
        'brand': 'tags',
        'Бренд': 'tags',

        # Pricing
        'price': 'price',
        'Цена': 'price',
        'old_price': 'original_price',
        'Старая цена': 'original_price',
        'marketing_price': 'marketing_price',

        # Stock
        'stock': 'stock_quantity',
        'Остаток': 'stock_quantity',
        'Доступное количество': 'stock_quantity',

        # Images
        'primary_image': 'main_image_url',
        'Главное изображение': 'main_image_url',
        'images': 'gallery_urls',
        'Изображения': 'gallery_urls',
        'Ссылки на изображения': 'gallery_urls',

        # Variants
        'Размер': 'size',
        'Цвет': 'color',
        'color': 'color',
        'size': 'size',

        # Status
        'visibility': 'status',
        'visible': 'status',
    }

    def _detect_file_type(self, file_path: str) -> str:
        """Detect file type from extension."""
        ext = file_path.lower().rsplit('.', 1)[-1]
        return ext

    def get_columns(self, file_path: str) -> list[str]:
        file_type = self._detect_file_type(file_path)

        if file_type == 'csv':
            return self._get_csv_columns(file_path)
        elif file_type == 'xlsx':
            return self._get_xlsx_columns(file_path)
        elif file_type == 'xls' and XLRD_AVAILABLE:
            return self._get_xls_columns(file_path)
        else:
            # Try xlsx as fallback
            return self._get_xlsx_columns(file_path)

    def _get_csv_columns(self, file_path: str) -> list[str]:
        """Get columns from CSV file."""
        encodings = ['utf-8', 'utf-8-sig', 'cp1251', 'latin1']

        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc, newline='') as f:
                    sample = f.read(4096)
                    f.seek(0)

                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                    reader = csv.reader(f, dialect)
                    headers = next(reader, [])

                    if headers:
                        return [h.strip() for h in headers]
            except (UnicodeDecodeError, csv.Error):
                continue

        return []

    def _get_xlsx_columns(self, file_path: str) -> list[str]:
        """Get columns from XLSX file."""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()

            if first_row:
                return [str(cell).strip() if cell else f'Column_{i}'
                        for i, cell in enumerate(first_row)]
            return []
        except Exception:
            return []

    def _get_xls_columns(self, file_path: str) -> list[str]:
        """Get columns from XLS file (legacy Excel format)."""
        if not XLRD_AVAILABLE:
            return []

        try:
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)

            if ws.nrows > 0:
                headers = []
                for i in range(ws.ncols):
                    cell = ws.cell_value(0, i)
                    headers.append(str(cell).strip() if cell else f'Column_{i}')
                return headers
            return []
        except Exception:
            return []

    def iter_rows(self, file_path: str) -> Iterator[dict[str, Any]]:
        file_type = self._detect_file_type(file_path)

        if file_type == 'csv':
            yield from self._iter_csv_rows(file_path)
        elif file_type == 'xlsx':
            yield from self._iter_xlsx_rows(file_path)
        elif file_type == 'xls' and XLRD_AVAILABLE:
            yield from self._iter_xls_rows(file_path)
        else:
            yield from self._iter_xlsx_rows(file_path)

    def _iter_csv_rows(self, file_path: str) -> Iterator[dict[str, Any]]:
        """Iterate CSV rows."""
        encodings = ['utf-8', 'utf-8-sig', 'cp1251', 'latin1']
        encoding = 'utf-8'

        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc, newline='') as f:
                    f.read(1024)
                    encoding = enc
                    break
            except UnicodeDecodeError:
                continue

        try:
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                sample = f.read(4096)
                f.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                reader = csv.DictReader(f, dialect=dialect)

                for row in reader:
                    yield {k.strip(): self.normalize_value(v) for k, v in row.items()}
        except Exception:
            return

    def _iter_xlsx_rows(self, file_path: str) -> Iterator[dict[str, Any]]:
        """Iterate XLSX rows."""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None)

            if not headers:
                wb.close()
                return

            headers = [str(h).strip() if h else f'Column_{i}'
                      for i, h in enumerate(headers)]

            for row in rows:
                if row and any(cell is not None for cell in row):
                    yield {
                        headers[i]: self.normalize_value(cell)
                        for i, cell in enumerate(row)
                        if i < len(headers)
                    }

            wb.close()
        except Exception:
            return

    def _iter_xls_rows(self, file_path: str) -> Iterator[dict[str, Any]]:
        """Iterate XLS rows."""
        if not XLRD_AVAILABLE:
            return

        try:
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)

            if ws.nrows < 2:
                return

            # Get headers
            headers = []
            for i in range(ws.ncols):
                cell = ws.cell_value(0, i)
                headers.append(str(cell).strip() if cell else f'Column_{i}')

            # Iterate rows
            for row_num in range(1, ws.nrows):
                row_data = {}
                for col_num in range(ws.ncols):
                    if col_num < len(headers):
                        value = ws.cell_value(row_num, col_num)
                        row_data[headers[col_num]] = self.normalize_value(value)

                if any(v is not None for v in row_data.values()):
                    yield row_data

        except Exception:
            return

    def get_suggested_mapping(self) -> dict[str, str]:
        """Return Ozon-specific column mappings."""
        return {k.lower(): v for k, v in self.KNOWN_COLUMNS.items()}

    def parse_visibility_status(self, value: Any) -> str:
        """Convert Ozon visibility to product status."""
        if value is None:
            return 'draft'

        value_str = str(value).lower().strip()

        if value_str in ('true', '1', 'да', 'yes', 'visible'):
            return 'published'
        elif value_str in ('false', '0', 'нет', 'no', 'hidden'):
            return 'draft'

        return 'draft'

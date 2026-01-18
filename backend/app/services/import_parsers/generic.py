"""
Generic CSV and XLSX parsers for user-defined formats.
"""

import csv
from typing import Any, Iterator

from openpyxl import load_workbook

from .base import BaseImportParser


class GenericCSVParser(BaseImportParser):
    """Parser for generic CSV files with any structure."""

    source_name = "CSV файл"
    supported_extensions = ['.csv']

    def __init__(self, delimiter: str = ',', encoding: str = 'utf-8'):
        self.delimiter = delimiter
        self.encoding = encoding

    def get_columns(self, file_path: str) -> list[str]:
        # Try different encodings
        encodings = [self.encoding, 'utf-8-sig', 'cp1251', 'latin1']

        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc, newline='') as f:
                    # Detect delimiter
                    sample = f.read(4096)
                    f.seek(0)

                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                    reader = csv.reader(f, dialect)
                    headers = next(reader, [])

                    if headers:
                        self.encoding = enc
                        self.delimiter = dialect.delimiter
                        return [h.strip() for h in headers]
            except (UnicodeDecodeError, csv.Error):
                continue

        return []

    def iter_rows(self, file_path: str) -> Iterator[dict[str, Any]]:
        columns = self.get_columns(file_path)
        if not columns:
            return

        with open(file_path, 'r', encoding=self.encoding, newline='') as f:
            reader = csv.DictReader(f, delimiter=self.delimiter)
            for row in reader:
                yield {k.strip(): self.normalize_value(v) for k, v in row.items()}

    def get_suggested_mapping(self) -> dict[str, str]:
        """Generic mapping - maps common column names."""
        return {
            # Russian
            'название': 'name',
            'наименование': 'name',
            'имя': 'name',
            'артикул': 'sku',
            'код': 'sku',
            'штрихкод': 'barcode',
            'штрих-код': 'barcode',
            'описание': 'short_description',
            'полное описание': 'long_description',
            'категория': 'category',
            'цена': 'price',
            'стоимость': 'price',
            'остаток': 'stock_quantity',
            'количество': 'stock_quantity',
            'изображение': 'main_image_url',
            'фото': 'main_image_url',
            'ссылка': 'external_url',
            'теги': 'tags',
            'метки': 'tags',
            # English
            'name': 'name',
            'title': 'name',
            'sku': 'sku',
            'article': 'sku',
            'barcode': 'barcode',
            'description': 'short_description',
            'category': 'category',
            'price': 'price',
            'stock': 'stock_quantity',
            'quantity': 'stock_quantity',
            'image': 'main_image_url',
            'photo': 'main_image_url',
            'url': 'external_url',
            'link': 'external_url',
            'tags': 'tags',
        }


class GenericXLSXParser(BaseImportParser):
    """Parser for generic Excel XLSX files."""

    source_name = "Excel файл"
    supported_extensions = ['.xlsx', '.xls']

    def __init__(self, sheet_name: str | None = None):
        self.sheet_name = sheet_name

    def get_columns(self, file_path: str) -> list[str]:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[self.sheet_name] if self.sheet_name else wb.active

            # Get first row as headers
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()

            if first_row:
                return [str(cell).strip() if cell else f'Column_{i}'
                        for i, cell in enumerate(first_row)]
            return []
        except Exception:
            return []

    def iter_rows(self, file_path: str) -> Iterator[dict[str, Any]]:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[self.sheet_name] if self.sheet_name else wb.active

            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None)

            if not headers:
                wb.close()
                return

            # Clean headers
            headers = [str(h).strip() if h else f'Column_{i}'
                      for i, h in enumerate(headers)]

            for row in rows:
                if row and any(cell is not None for cell in row):
                    yield {
                        headers[i]: self.normalize_value(cell)
                        for i, cell in enumerate(row)
                        if i < len(headers)
                    }

            wb.close()
        except Exception:
            return

    def get_suggested_mapping(self) -> dict[str, str]:
        """Same as CSV generic mapping."""
        return GenericCSVParser().get_suggested_mapping()
